import cantools
import pandas as pd
from openpyxl import load_workbook
import json
import os
import shutil

# ---------------------------------------------------
# USER CONFIGURATION
# ---------------------------------------------------

INPUT_TEMPLATE = "KGM/KGM_Template.xlsx"   # Knowledge Graph Matrix (KGM) template
OUTPUT_EXCEL = "KGM.xlsx"              # Output - KGM populated with DBC file entries

DBC_FILE = "DBC/boening.dbc"
UNIT_MAPPING_FILE = "KGM/unit_mapping.json"

CAN_PLATFORM = "can2"
SNIFFING_SENSOR = "can2_sniffer"


# ---------------------------------------------------
# QUDT UNIT MAPPING
# ---------------------------------------------------

def load_unit_mapping(json_path=UNIT_MAPPING_FILE):
    """Load unit mappings from JSON file."""
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"❌ Unit mapping file not found: {json_path}")
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


def map_to_qudt_unit(raw_unit: str, unit_map: dict) -> str:
    if not raw_unit:
        return ""
    key = raw_unit.strip().lower()
    mapped = unit_map.get(key)
    if mapped is None:
        print(f"⚠️ Warning: Unknown unit '{raw_unit}' — keeping as-is.")
        return raw_unit
    return mapped


# ---------------------------------------------------
# DBC PARSING
# ---------------------------------------------------

def parse_dbc(file_path, unit_map):
    db = cantools.database.load_file(file_path)

    signals_data = []
    messages_data = []
    encodings_data = []
    nodes_data = set()

    for msg in db.messages:

        transmitters = msg.senders if msg.senders else ["Unknown"]
        transmitters = ["Unknown" if t == "Vector__XXX" else t for t in transmitters]
        nodes_data.update(transmitters)

        signal_names = [sig.name for sig in msg.signals]
        encoded_via = [f"{sig.name}Encoding" for sig in msg.signals]

        messages_data.append({
            "Individual": msg.name,
            "rdf:type": "dbc:Message",
            "dbc:dataLength": msg.length,
            "dbc:encodedVia": ", ".join(encoded_via),
            "dbc:hasDecID": msg.frame_id,
            "dbc:hasSignal": ", ".join(signal_names),
            "dbc:hasTransmitter": ", ".join(transmitters),
            "sosa:isObservedBy": SNIFFING_SENSOR
        })

        for sig in msg.signals:

            receivers = sig.receivers if sig.receivers else ["Unknown"]
            receivers = ["Unknown" if r == "Vector__XXX" else r for r in receivers]

            unit = map_to_qudt_unit(sig.unit or "", unit_map)

            nodes_data.update(receivers)

            # SIGNAL ENTRY
            signals_data.append({
                "Individual": sig.name,
                "rdf:type": "dbc:Signal",
                "dbc:decodedVia": f"{sig.name}Encoding",
                "dbc:hasReceiver": ", ".join(receivers),
                "dbc:isPartOf": msg.name,
                "qudt:hasUnit": unit,
                "sosa:isObservedBy": SNIFFING_SENSOR
            })

            # ENCODING ENTRY
            encodings_data.append({
                "Individual": f"{sig.name}Encoding",
                "rdf:type": "dbc:SignalEncoding",
                "dbc:bitLenght": sig.length,
                "dbc:bitStart": sig.start,
                "dbc:signed": str(sig.is_signed).lower(),
                "qudt:byteOrder": "LittleEndian" if sig.byte_order == "little_endian" else "BigEndian",
                "qudt:conversionMultiplier": sig.scale if sig.scale is not None else 1,
                "qudt:conversionOffset": sig.offset if sig.offset is not None else 0,
                "qudt:maxInclusive": sig.maximum if sig.maximum is not None else "",
                "qudt:minInclusive": sig.minimum if sig.minimum is not None else ""
            })

    # NODE ENTRIES
    node_list = [{"Individual": n, "rdf:type": "dbc:Node"} for n in sorted(nodes_data)]

    return signals_data, messages_data, encodings_data, node_list


# ---------------------------------------------------
# STATIC TABLES
# ---------------------------------------------------

def get_platform_data():
    return [{
        "Individual": CAN_PLATFORM,
        "rdf:type": "sosa:Platform",
        "sosa:hosts": SNIFFING_SENSOR
    }]


def get_sensor_data(observed_messages: list):
    observed_str = ", ".join(observed_messages)
    return [{
        "Individual": SNIFFING_SENSOR,
        "rdf:type": "sosa:Sensor",
        "sosa:isHostedBy": CAN_PLATFORM,
        "sosa:madeObservation": "NA",
        "sosa:observes": observed_str
    }]


# ---------------------------------------------------
# KGM EXCEL WRITING
# ---------------------------------------------------

def append_to_existing_sheet(data, excel_path, sheet_name, columns):
    df = pd.DataFrame(data, columns=columns)

    book = load_workbook(excel_path)
    if sheet_name not in book.sheetnames:
        raise ValueError(f"❌ Sheet '{sheet_name}' not found in '{excel_path}'")

    sheet = book[sheet_name]
    start_row = sheet.max_row + 1

    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            sheet.cell(row=r_idx, column=c_idx, value=value)

    book.save(excel_path)
    print(f"✅ Appended {len(df)} rows to '{excel_path}' → sheet '{sheet_name}', starting at row {start_row}")


# ---------------------------------------------------
# MAIN SCRIPT
# ---------------------------------------------------

if __name__ == "__main__":

    # Copy the template to output
    if not os.path.exists(INPUT_TEMPLATE):
        raise FileNotFoundError(f"❌ Template Excel file not found: {INPUT_TEMPLATE}")

    shutil.copyfile(INPUT_TEMPLATE, OUTPUT_EXCEL)
    print(f"📄 Created working Excel file: {OUTPUT_EXCEL}\n")

    # Load mapping
    unit_map = load_unit_mapping(UNIT_MAPPING_FILE)

    # Parse DBC
    signals, messages, encodings, nodes = parse_dbc(DBC_FILE, unit_map)

    # Static sets
    platforms = get_platform_data()
    sensors = get_sensor_data([m["Individual"] for m in messages])

    # Append each dataset
    append_to_existing_sheet(signals, OUTPUT_EXCEL, "Signal",
        ["Individual", "rdf:type", "dbc:decodedVia", "dbc:hasReceiver", "dbc:isPartOf",
         "qudt:hasUnit", "sosa:isObservedBy"]
    )

    append_to_existing_sheet(messages, OUTPUT_EXCEL, "Message",
        ["Individual", "rdf:type", "dbc:dataLength", "dbc:encodedVia", "dbc:hasDecID",
         "dbc:hasSignal", "dbc:hasTransmitter", "sosa:isObservedBy"]
    )

    append_to_existing_sheet(encodings, OUTPUT_EXCEL, "SignalEncoding",
        ["Individual", "rdf:type", "dbc:bitLenght", "dbc:bitStart", "dbc:signed",
         "qudt:byteOrder", "qudt:conversionMultiplier", "qudt:conversionOffset",
         "qudt:maxInclusive", "qudt:minInclusive"]
    )

    append_to_existing_sheet(platforms, OUTPUT_EXCEL, "Platform",
        ["Individual", "rdf:type", "sosa:hosts"]
    )

    append_to_existing_sheet(nodes, OUTPUT_EXCEL, "Node",
        ["Individual", "rdf:type"]
    )

    append_to_existing_sheet(sensors, OUTPUT_EXCEL, "Sensor",
        ["Individual", "rdf:type", "sosa:isHostedBy", "sosa:madeObservation", "sosa:observes"]
    )

    print("\n🎉 KGM.xlsx generated successfully!")

